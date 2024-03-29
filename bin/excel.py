# -*- coding: utf-8 -*-
"""
Created on Fri Sep 27 10:01:22 2019

@author: fener_000
"""
from openpyxl import Workbook
class excel():
    def __init__(self,title):
        self.file = Workbook() 
        self.ws = self.file.active
        self.ws.title = str(title)
        
    def write(self,):
        pass
from openpyxl import load_workbook
import csv
def update_xlsx(src, dest):
    #Open an xlsx for reading
    wb = load_workbook(filename = sample.xlsx)
    #Get the current Active Sheet
    ws = wb.get_active_sheet()
    #You can also select a particular sheet
    #based on sheet name
    #ws = wb.get_sheet_by_name("Sheet1")
    #Open the csv file
    with open(src) as fin:
        #read the csv
        reader = csv.reader(fin)
        #enumerate the rows, so that you can
        #get the row index for the xlsx
        for index,row in enumerate(reader):
            #Assuming space separated,
            #Split the row to cells (column)
            row = row[0].split()
            #Access the particular cell and assign
            #the value from the csv row
            ws.cell(row=index,column=7).value = row[2]
            ws.cell(row=index,column=8).value = row[3]            
    #save the csb file
    wb.save(sample.xlsx)

