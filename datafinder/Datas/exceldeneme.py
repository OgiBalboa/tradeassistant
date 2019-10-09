"""
Hos Geldin !
date = 02.10.2019
update = 02.10.2019
author = @ogibalboa
---------------------
Excelden bilgi alip ayiracak ve yeni bir excel dosyasina yazacak program
dosya adı yazımı : all_hotmelt20190207_034970_190207104054p.csv
"""
from openpyxl import Workbook, load_workbook
def exp():
    wb = Workbook()
    ws = wb.active
def imp():
    global sr
    global wb
    wb = load_workbook(filename = 'all_hotmelt20190207_034970_190207104054p.csv')
    sr = wb['Sheet']

def yazdir():
    ws['A1'] = 42
    
def oku():
    print(sr['A2'].value)
    
print("Hos geldin")

imp()
oku()

wb.save("all_hotmelt20190207_034970_190207104054p.csv")
